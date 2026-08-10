from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from ..models import NormalizedUnit, ParsedDocument
from .base import DocumentParser


class XlsxParser(DocumentParser):
    def parse(
        self,
        data: bytes,
        filename: str,
        mime_type: str | None,
    ) -> ParsedDocument:
        workbook = load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
        )

        units: list[NormalizedUnit] = []

        try:
            for sheet_index, worksheet in enumerate(
                workbook.worksheets,
                start=1,
            ):
                rows: list[str] = []

                for row in worksheet.iter_rows(
                    values_only=True,
                ):
                    values = [
                        "" if value is None else str(value)
                        for value in row
                    ]

                    if any(
                        value.strip()
                        for value in values
                    ):
                        rows.append(
                            " | ".join(values)
                        )

                text = "\n".join(rows)

                units.append(
                    NormalizedUnit(
                        index=sheet_index,
                        kind="sheet",
                        label=worksheet.title,
                        text=text,
                        metadata={
                            "sheet_name":
                                worksheet.title,
                            "row_count": len(rows),
                        },
                    )
                )

            full_text = "\n\n".join(
                unit.text
                for unit in units
                if unit.text
            )

            return ParsedDocument(
                filename=filename,
                extension=Path(
                    filename,
                ).suffix.lower(),
                mime_type=mime_type,
                parser="openpyxl",
                units=units,
                full_text=full_text,
                metadata={
                    "sheet_count": len(units),
                },
            )

        finally:
            workbook.close()